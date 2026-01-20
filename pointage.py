import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import datetime


def create_pointage_file():
    # Créer un nouveau classeur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pointage"

    # Créer une feuille pour le récapitulatif
    ws_recap = wb.create_sheet("Récapitulatif")

    # ========== FEUILLE POINTAGE ==========

    # Entêtes
    headers = ["DATE", "ENTRE", "PAUSE DEBUT", "PAUSE FIN", "SORTIE",
               "TYPE JOUR", "HEURES", "HEURES SUP", "RETARD", "STATUT", "COMMENTAIRE"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Données d'exemple avec formules intégrées
    sample_data = [
        # Date, Entre, Pause Début, Pause Fin, Sortie, Type Jour, Heures, Heures Sup, Retard, Statut, Commentaire
        ["01/01/2024", "8:21", "12:59", "13:59", "17:59",
         '=SI(OU(B2="Congé";B2="Férié");SI(B2="Congé";"Congé";"Férié");"Normal")',
         '=SI(OU(F2="Congé";F2="Férié");0;SI(ET(B2>0;E2>0);(E2-B2)-(D2-C2);0))',
         '=SI(F2="Normal";MAX(0;G2-TEMPS(8;0;0));0)',
         '=SI(ET(F2="Normal";B2>TEMPS(8;30;0));B2-TEMPS(8;30;0);0)',
         '=SI(F2="Congé";"Congé Payé";SI(F2="Férié";"Férié";SI(G2<TEMPS(7;30;0);"Incomplet";SI(I2>0;"Avec Retard";"Normal"))))',
         ""],

        ["02/01/2024", "Congé", "", "", "",
         '=SI(OU(B3="Congé";B3="Férié");SI(B3="Congé";"Congé";"Férié");"Normal")',
         '=SI(OU(F3="Congé";F3="Férié");0;SI(ET(B3>0;E3>0);(E3-B3)-(D3-C3);0))',
         '=SI(F3="Normal";MAX(0;G3-TEMPS(8;0;0));0)',
         '=SI(ET(F3="Normal";B3>TEMPS(8;30;0));B3-TEMPS(8;30;0);0)',
         '=SI(F3="Congé";"Congé Payé";SI(F3="Férié";"Férié";SI(G3<TEMPS(7;30;0);"Incomplet";SI(I3>0;"Avec Retard";"Normal"))))',
         "RTT"],

        ["03/01/2024", "8:45", "12:30", "13:30", "18:00",
         '=SI(OU(B4="Congé";B4="Férié");SI(B4="Congé";"Congé";"Férié");"Normal")',
         '=SI(OU(F4="Congé";F4="Férié");0;SI(ET(B4>0;E4>0);(E4-B4)-(D4-C4);0))',
         '=SI(F4="Normal";MAX(0;G4-TEMPS(8;0;0));0)',
         '=SI(ET(F4="Normal";B4>TEMPS(8;30;0));B4-TEMPS(8;30;0);0)',
         '=SI(F4="Congé";"Congé Payé";SI(F4="Férié";"Férié";SI(G4<TEMPS(7;30;0);"Incomplet";SI(I4>0;"Avec Retard";"Normal"))))',
         "Embouteillage"],

        ["04/01/2024", "7:45", "12:00", "13:00", "16:30",
         '=SI(OU(B5="Congé";B5="Férié");SI(B5="Congé";"Congé";"Férié");"Normal")',
         '=SI(OU(F5="Congé";F5="Férié");0;SI(ET(B5>0;E5>0);(E5-B5)-(D5-C5);0))',
         '=SI(F5="Normal";MAX(0;G5-TEMPS(8;0;0));0)',
         '=SI(ET(F5="Normal";B5>TEMPS(8;30;0));B5-TEMPS(8;30;0);0)',
         '=SI(F5="Congé";"Congé Payé";SI(F5="Férié";"Férié";SI(G5<TEMPS(7;30;0);"Incomplet";SI(I5>0;"Avec Retard";"Normal"))))',
         ""],

        ["05/01/2024", "Férié", "", "", "",
         '=SI(OU(B6="Congé";B6="Férié");SI(B6="Congé";"Congé";"Férié");"Normal")',
         '=SI(OU(F6="Congé";F6="Férié");0;SI(ET(B6>0;E6>0);(E6-B6)-(D6-C6);0))',
         '=SI(F6="Normal";MAX(0;G6-TEMPS(8;0;0));0)',
         '=SI(ET(F6="Normal";B6>TEMPS(8;30;0));B6-TEMPS(8;30;0);0)',
         '=SI(F6="Congé";"Congé Payé";SI(F6="Férié";"Férié";SI(G6<TEMPS(7;30;0);"Incomplet";SI(I6>0;"Avec Retard";"Normal"))))',
         "Jour de l'an"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajouter une ligne vierge avec formules
    for row in range(7, 10):
        # Date (hier + incrément)
        ws.cell(row=row, column=1, value=f'=A{row - 1}+1')

        # Formules pour les autres colonnes
        ws.cell(row=row,
                column=6).value = f'=SI(OU(B{row}="Congé";B{row}="Férié");SI(B{row}="Congé";"Congé";"Férié");"Normal")'
        ws.cell(row=row,
                column=7).value = f'=SI(OU(F{row}="Congé";F{row}="Férié");0;SI(ET(B{row}>0;E{row}>0);(E{row}-B{row})-(D{row}-C{row});0))'
        ws.cell(row=row, column=8).value = f'=SI(F{row}="Normal";MAX(0;G{row}-TEMPS(8;0;0));0)'
        ws.cell(row=row, column=9).value = f'=SI(ET(F{row}="Normal";B{row}>TEMPS(8;30;0));B{row}-TEMPS(8;30;0);0)'
        ws.cell(row=row,
                column=10).value = f'=SI(F{row}="Congé";"Congé Payé";SI(F{row}="Férié";"Férié";SI(G{row}<TEMPS(7;30;0);"Incomplet";SI(I{row}>0;"Avec Retard";"Normal"))))'

    # Formatage des colonnes
    # Format Date
    for cell in ws["A"]:
        if cell.row > 1:  # Ne pas formater l'entête
            cell.number_format = "DD/MM/YYYY"

    # Format Heure pour les colonnes B, C, D, E
    for col in ["B", "C", "D", "E"]:
        for cell in ws[col]:
            if cell.row > 1 and cell.value and not isinstance(cell.value, str):
                cell.number_format = "HH:MM"

    # Format Heures avec [h]:mm pour >24h
    for col in ["G", "H", "I"]:
        for cell in ws[col]:
            if cell.row > 1:
                cell.number_format = "[h]:mm"

    # Mise en forme manuelle (simule le formatage conditionnel)
    for row in range(2, 11):
        type_jour = ws.cell(row=row, column=6).value

        # Jours spéciaux en gris
        if type_jour and isinstance(type_jour, str) and type_jour in ["Congé", "Férié"]:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                                                                fill_type="solid")

    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 12, 'B': 8, 'C': 12, 'D': 10, 'E': 8,
        'F': 10, 'G': 10, 'H': 10, 'I': 8, 'J': 12, 'K': 20
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ========== FEUILLE RÉCAPITULATIF ==========

    # Titre
    ws_recap["A1"] = "RÉCAPITULATIF MENSUEL"
    ws_recap["A1"].font = Font(bold=True, size=14)
    ws_recap.merge_cells("A1:B1")

    # Données du récapitulatif
    recap_data = [
        ["TOTAL HEURES TRAVAILLÉES", '=SOMME(Pointage!G:G)'],
        ["TOTAL HEURES SUP", '=SOMME(Pointage!H:H)'],
        ["TOTAL RETARD", '=SOMME(Pointage!I:I)'],
        ["JOURS TRAVAILLÉS", '=NB.SI(Pointage!F:F;"Normal")'],
        ["JOURS DE CONGÉ", '=NB.SI(Pointage!F:F;"Congé")'],
        ["JOURS FÉRIÉS", '=NB.SI(Pointage!F:F;"Férié")'],
        ["TAUX D'ABSENTÉISME", '=E5/(E4+E5+E6)']
    ]

    for i, (label, formula) in enumerate(recap_data, 3):
        ws_recap.cell(row=i, column=1, value=label)
        ws_recap.cell(row=i, column=2, value=formula)
        ws_recap.cell(row=i, column=1).font = Font(bold=True)

    # Formater les cellules de résultats
    for row in range(3, 10):
        ws_recap.cell(row=row, column=2).number_format = '[h]:mm'

    ws_recap.column_dimensions['A'].width = 25
    ws_recap.column_dimensions['B'].width = 15

    # Ajouter des bordures
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in range(3, 10):
        for col in range(1, 3):
            ws_recap.cell(row=row, column=col).border = thin_border

    # Sauvegarder le fichier
    filename = "Pointage_Simple_DIVA.xlsx"
    wb.save(filename)
    print(f"✅ Fichier créé avec succès : {filename}")
    print("📋 Le fichier contient :")
    print("   - Toutes les formules françaises")
    print("   - Données d'exemple")
    print("   - Récapitulatif automatique")
    print("   - Mise en forme de base")
    print("\n⚠️ Pour ajouter le formatage conditionnel :")
    print("   1. Ouvrez le fichier dans Excel")
    print("   2. Sélectionnez A2:J10")
    print("   3. Mise en forme conditionnelle > Nouvelle règle")
    print("   4. Utilisez les formules mentionnées précédemment")


# Exécuter le script
if __name__ == "__main__":
    try:
        create_pointage_file()
        print("\n✅ Fichier prêt ! Ouvrez-le dans Excel.")
    except Exception as e:
        print(f"Erreur : {e}")